#!/usr/bin/env python3
"""
validate_compliance.py — Gate COMPLIANCE (checks mécaniques)
Usage: python validate_compliance.py <fichier.xlsx> [--sheet STOCK_MAITRE] [--pool POOL_ID]

Checks (zéro token) :
  B1.1 Réponse absente
  B1.2 Réponse dans le libellé
  B1.3 Libellé > 10 mots
  B1.4 Formulation négative complexe
  F5   Qualificatifs éditoriaux
  HOM  Homonymes ambigus (liste configurable)
  VEI  Marqueurs veille (non bloquant)

Retourne : GO / NO_GO + rapport détaillé
"""

import sys
import re
import argparse
import openpyxl
from collections import defaultdict

# ── CONFIG ─────────────────────────────────────────────────────────────────

QUALIFICATIFS = [
    "célèbre", "légendaire", "fameux", "illustre", "réputé", "renommé",
    "incontournable", "emblématique", "mythique", "iconique"
]

HOMONYMES = [
    # Ajouter selon le thème du quiz
    # Exemple CDM : "ronaldo", "morales"
    # Mayenne : aucun par défaut
]

NEGATIFS_COMPLEXES = [
    r"\bni\s+\w+\s+ni\b",
    r"\bsans\s+que\b",
    r"\bexcepté\b",
    r"\bhormis\b",
    r"\bà\s+l[''']exception\b",
    r"\bà\s+moins\s+que\b",
]

MARQUEURS_VEILLE = [
    "dernier", "dernière", "premier", "première",
    "jamais", "encore jamais", "seul", "unique",
    "record", "recordman", "recordwoman",
    "plus grand", "plus petit", "plus rapide",
    "meilleur", "pire", "à ce jour",
]

# Mapping colonnes selon format détecté
FORMATS = {
    "STOCK_MAITRE": {"q_id": "retroId", "question": "question", "answer": "answer", "pool": "poolHistorique", "niveau": "niveauOrigine"},
    "B5_TABLEUR":   {"q_id": "ID Global", "question": "Question", "answer": "Bonne réponse", "pool": "Pool principal", "niveau": "Difficulté cible pool"},
    "QUESTIONS":    {"q_id": "Q_ID", "question": "LIBELLÉ", "answer": "RÉPONSE", "pool": "POOL_ID", "niveau": "CIBLE_NIVEAU"},
}

# ── LECTURE XLSX ────────────────────────────────────────────────────────────

def detect_format(headers):
    for fmt_name, mapping in FORMATS.items():
        if mapping["question"] in headers and mapping["answer"] in headers:
            return fmt_name, mapping
    return None, None

def load_questions(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        sheets_to_try = [sheet_name]
    else:
        sheets_to_try = wb.sheetnames

    questions = []
    for sname in sheets_to_try:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Chercher la ligne d'en-tête (row 1 ou 3 selon format B5)
        headers = None
        header_row_idx = None
        for i, row in enumerate(rows[:5]):
            row_vals = [str(v) if v is not None else "" for v in row]
            fmt_name, mapping = detect_format(row_vals)
            if mapping:
                headers = row_vals
                header_row_idx = i
                break

        if not headers:
            continue

        col = {v: headers.index(v) for v in mapping.values() if v in headers}

        for row in rows[header_row_idx + 1:]:
            if not any(v is not None for v in row):
                continue
            def get(field):
                key = mapping.get(field)
                if key and key in col:
                    v = row[col[key]]
                    return str(v).strip() if v is not None else ""
                return ""

            questions.append({
                "q_id":     get("q_id") or f"row_{len(questions)+1}",
                "question": get("question"),
                "answer":   get("answer"),
                "pool":     get("pool"),
                "niveau":   get("niveau"),
                "sheet":    sname,
            })

    wb.close()
    return questions

# ── CHECKS ──────────────────────────────────────────────────────────────────

def check_question(q):
    fails = []
    warnings = []
    libelle = q["question"].strip()
    reponse = q["answer"].strip()
    libelle_lower = libelle.lower()

    # B1.1 — Réponse absente
    if not reponse:
        fails.append("B1.1_RÉPONSE_ABSENTE")

    # B1.2 — Réponse dans le libellé
    if reponse and reponse.lower() in libelle_lower:
        fails.append(f"B1.2_RÉPONSE_DANS_LIBELLÉ ('{reponse}' détecté)")

    # B1.3 — Libellé > 10 mots
    if libelle:
        nb_mots = len(libelle.split())
        if nb_mots > 10:
            fails.append(f"B1.3_LIBELLÉ_TROP_LONG ({nb_mots} mots)")

    # B1.4 — Formulation négative complexe
    for pattern in NEGATIFS_COMPLEXES:
        if re.search(pattern, libelle_lower):
            fails.append(f"B1.4_FORMULATION_NÉGATIVE_COMPLEXE ('{pattern}')")
            break

    # F5 — Qualificatifs éditoriaux
    for q_word in QUALIFICATIFS:
        if re.search(r"\b" + q_word + r"\b", libelle_lower):
            fails.append(f"F5_QUALIFICATIF_ÉDITORIAL ('{q_word}')")
            break

    # HOM — Homonymes ambigus
    for hom in HOMONYMES:
        if re.search(r"\b" + hom.lower() + r"\b", libelle_lower):
            fails.append(f"HOM_HOMONYME_AMBIGU ('{hom}')")

    # VEI — Marqueurs veille (warning seulement)
    for m in MARQUEURS_VEILLE:
        if m in libelle_lower:
            warnings.append(f"VEI_MARQUEUR_VEILLE ('{m}')")
            break

    return fails, warnings

# ── RAPPORT ──────────────────────────────────────────────────────────────────

def run(path, sheet_name=None, pool_filter=None):
    print(f"\n{'='*60}")
    print(f"COMPLIANCE REPORT — {path}")
    if sheet_name:
        print(f"Sheet : {sheet_name}")
    if pool_filter:
        print(f"Pool  : {pool_filter}")
    print(f"{'='*60}\n")

    questions = load_questions(path, sheet_name)
    if pool_filter:
        questions = [q for q in questions if q["pool"] == pool_filter]

    if not questions:
        print("⚠️  Aucune question trouvée. Vérifier le fichier et le format.")
        sys.exit(1)

    print(f"Questions chargées : {len(questions)}\n")

    total_go = 0
    total_fail = 0
    fail_ids = []
    warn_count = 0

    for q in questions:
        fails, warnings = check_question(q)
        status = "FAIL" if fails else "GO"
        if fails:
            total_fail += 1
            fail_ids.append(q["q_id"])
            print(f"[{status}] {q['q_id']} (Pool: {q['pool'] or '?'})")
            for f in fails:
                print(f"         ✗ {f}")
            for w in warnings:
                print(f"         ⚠ {w}")
        else:
            total_go += 1
            if warnings:
                warn_count += 1
                print(f"[{status}] {q['q_id']} (Pool: {q['pool'] or '?'})")
                for w in warnings:
                    print(f"         ⚠ {w}")

    print(f"\n{'─'*60}")
    print(f"RÉSUMÉ")
    print(f"  Total     : {len(questions)}")
    print(f"  GO        : {total_go}")
    print(f"  FAIL      : {total_fail}")
    print(f"  WARNINGS  : {warn_count}")

    verdict = "NO_GO" if total_fail > 0 else "GO"
    print(f"\nCOMPLIANCE_GLOBAL : {verdict}")
    if fail_ids:
        print(f"Q_IDS_FAIL : {fail_ids}")
    print(f"{'='*60}\n")

    return verdict

# ── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Validation COMPLIANCE — FACTORY")
    parser.add_argument("fichier", help="Chemin vers le fichier xlsx")
    parser.add_argument("--sheet", default=None, help="Nom de la feuille (défaut: auto-detect)")
    parser.add_argument("--pool", default=None, help="Filtrer sur un pool spécifique")
    args = parser.parse_args()

    verdict = run(args.fichier, args.sheet, args.pool)
    sys.exit(0 if verdict == "GO" else 1)
