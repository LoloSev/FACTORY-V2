"""
generate_xlsx_view.py
ROLE    : Generate human-readable xlsx view from TSV source files.
INPUT   : _FACTORY/_LIGNES/[THEME]/ (TSV + yaml)
OUTPUT  : _FACTORY/_LIGNES/[THEME]/VIEW_[THEME].xlsx
TRIGGER : On demand / at HUMAN_GATE
SOURCE  : TSV files are the source of truth. xlsx is read-only for humans.
"""

import sys
import csv
import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# ── Style constants ──────────────────────────────────────────────────────────
HEADER_FONT  = Font(bold=True, color='FFFFFF', size=10)
HEADER_FILL  = PatternFill('solid', fgColor='2F4F8F')
WARN_FILL    = PatternFill('solid', fgColor='FFA500')
FAIL_FILL    = PatternFill('solid', fgColor='CC0000')
PASS_FILL    = PatternFill('solid', fgColor='228B22')
ALIGN_CTR    = Alignment(horizontal='center')

STATUS_FILLS = {
    'FAIL': FAIL_FILL, 'WARN': WARN_FILL, 'PASS': PASS_FILL,
    'READY_B3': PASS_FILL, 'REWRITE': WARN_FILL, 'DRAFT': None,
    'READY_EXPORT': PASS_FILL, 'DROP': FAIL_FILL,
}

SHEETS_ORDER = ['CONFIG', 'ITEMS', 'ANGLES', 'POOLS', 'QUESTIONS', 'DISTRACTEURS', 'QA']
TSV_MAP = {
    'ITEMS': 'ITEMS.tsv', 'ANGLES': 'ANGLES.tsv', 'POOLS': 'POOLS.tsv',
    'QUESTIONS': 'QUESTIONS.tsv', 'DISTRACTEURS': 'DISTRACTEURS.tsv', 'QA': 'QA.tsv',
}
STATUS_COLS = {'QA_STATUS', 'STATUT_B2', 'STATUT_B3', 'COUVERTURE_NIVEAU',
               'FAISABILITE', 'DECISION_RUNTIME', 'NIVEAU_QUESTION',
               'NIVEAU_ANGLE', 'NIVEAU_POTENTIEL', 'NIVEAU_CONFIRME'}


def style_header(ws, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CTR


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


def apply_status_fill(cell):
    val = str(cell.value or '').strip().upper()
    fill = STATUS_FILLS.get(val)
    if fill:
        cell.fill = fill
        cell.font = Font(bold=True, color='FFFFFF', size=10)


def write_tsv_sheet(wb, sheet_name, tsv_path):
    ws = wb.create_sheet(sheet_name)
    if not tsv_path.exists():
        ws.cell(row=1, column=1, value=f'MISSING: {tsv_path.name}')
        return
    with open(tsv_path, encoding='utf-8', newline='') as f:
        reader = csv.reader(f, delimiter='\t')
        headers = None
        for r, row in enumerate(reader, 1):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                if r == 1:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = ALIGN_CTR
                elif headers and c <= len(headers) and headers[c-1] in STATUS_COLS:
                    apply_status_fill(cell)
            if r == 1:
                headers = row
    auto_width(ws)
    ws.freeze_panes = 'A2'


def write_config_sheet(wb, yaml_path):
    ws = wb.create_sheet('CONFIG', 0)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    if not yaml_path.exists():
        ws.cell(row=1, column=1, value='MISSING: CONFIG.yaml')
        return
    with open(yaml_path, encoding='utf-8') as f:
        data = yaml.safe_load(f)
    headers = ['CHAMP', 'VALEUR']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CTR
    for r, (k, v) in enumerate(data.items(), 2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=str(v))


def generate(theme_dir: Path):
    theme = theme_dir.name
    out_path = theme_dir / f'VIEW_{theme}.xlsx'

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    write_config_sheet(wb, theme_dir / 'CONFIG.yaml')
    for sheet_name in SHEETS_ORDER[1:]:  # skip CONFIG already done
        tsv_name = TSV_MAP.get(sheet_name)
        if tsv_name:
            write_tsv_sheet(wb, sheet_name, theme_dir / tsv_name)

    wb.save(out_path)
    print(f'VIEW generated: {out_path}')
    return out_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python generate_xlsx_view.py [THEME_DIR]')
        print('Example: python generate_xlsx_view.py ../_LIGNES/MAYENNE')
        sys.exit(1)
    theme_dir = Path(sys.argv[1])
    if not theme_dir.is_dir():
        print(f'ERROR: directory not found: {theme_dir}')
        sys.exit(1)
    generate(theme_dir)
