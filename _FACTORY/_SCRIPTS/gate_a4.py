#!/usr/bin/env python3
"""
gate_a4.py — Gate A4 : validation architecture pools
Usage: python gate_a4.py <A4_LIGNE.xlsx> <LIGNE>

Checks (RULE-ARCH-001/002/003/004) :
  ARCH-1  Total pools = 20
  ARCH-2  Composition : 2 IF-SF + 3 IF-ROT + 15 QV
  ARCH-3  Tous les pools ont CIBLE_NIVEAU renseigné
  ARCH-4  Tous les pools ont STOCK_CIBLE > 0
  ARCH-5  Tous les pools ont THÈME_ÉDITORIAL renseigné
  ARCH-6  Crescendo N1/N2/N3 cohérent sur les QV

GO   → déverrouille B2 dans pipeline_state.json
NO_GO → pose flag B2_LOCKED + met à jour dashboard
"""

import sys
import json
import os
import argparse
import openpyxl
from pathlib import Path
from datetime import datetime

ROOT         = Path(__file__).parent.parent
STATE_FILE   = ROOT / "_STATE" / "pipeline_state.json"
DASH_FILE    = ROOT / "_STATE" / "DASHBOARD_STATE.json"

# Crescendo attendu sur les QV (position → niveau)
# Q1-Q5 = N1, Q6-Q15 = N2, Q16-Q20 = N3
def niveau_attendu(position_str):
    try:
        pos = int(position_str.replace("Q", ""))
    except Exception:
        return None
    if 1 <= pos <= 5:
        return "N1"
    elif 6 <= pos <= 15:
        return "N2"
    elif 16 <= pos <= 20:
        return "N3"
    return None

# ── LECTURE A4 ───────────────────────────────────────────────────────────────

def load_pools(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "POOLS" not in wb.sheetnames:
        wb.close()
        return None, "Feuille POOLS absente"
    ws = wb["POOLS"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return None, "Feuille POOLS vide"

    headers = [str(v) if v is not None else "" for v in rows[0]]
    col = {h: i for i, h in enumerate(headers)}

    required = ["POOL_ID", "TYPE", "POSITION_QUIZ", "CIBLE_NIVEAU", "THÈME_ÉDITORIAL", "STOCK_CIBLE"]
    missing = [r for r in required if r not in col]
    if missing:
        return None, f"Colonnes manquantes : {missing}"

    pools = []
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        def get(field):
            idx = col.get(field)
            if idx is not None and idx < len(row):
                v = row[idx]
                return str(v).strip() if v is not None else ""
            return ""
        pools.append({
            "pool_id":   get("POOL_ID"),
            "type":      get("TYPE"),
            "position":  get("POSITION_QUIZ"),
            "niveau":    get("CIBLE_NIVEAU"),
            "theme":     get("THÈME_ÉDITORIAL"),
            "stock":     get("STOCK_CIBLE"),
        })

    return pools, None

# ── CHECKS ───────────────────────────────────────────────────────────────────

def run_checks(pools):
    fails = []
    warnings = []

    # ARCH-1 : total = 20
    if len(pools) != 20:
        fails.append(f"ARCH-1 : {len(pools)} pools détectés — attendu 20")

    # ARCH-2 : composition
    types = [p["type"] for p in pools]
    n_ifsf  = types.count("IF-SF")
    n_ifrot = types.count("IF-ROT")
    n_qv    = types.count("QV")
    if n_ifsf != 2:
        fails.append(f"ARCH-2 : {n_ifsf} IF-SF — attendu 2")
    if n_ifrot != 3:
        fails.append(f"ARCH-2 : {n_ifrot} IF-ROT — attendu 3")
    if n_qv != 15:
        fails.append(f"ARCH-2 : {n_qv} QV — attendu 15")

    # ARCH-3 : CIBLE_NIVEAU renseigné
    sans_niveau = [p["pool_id"] for p in pools if not p["niveau"]]
    if sans_niveau:
        fails.append(f"ARCH-3 : CIBLE_NIVEAU absent sur {sans_niveau}")

    # ARCH-4 : STOCK_CIBLE > 0
    sans_stock = []
    for p in pools:
        try:
            if float(p["stock"]) <= 0:
                sans_stock.append(p["pool_id"])
        except (ValueError, TypeError):
            sans_stock.append(p["pool_id"])
    if sans_stock:
        fails.append(f"ARCH-4 : STOCK_CIBLE manquant/nul sur {sans_stock}")

    # ARCH-5 : THÈME_ÉDITORIAL renseigné
    sans_theme = [p["pool_id"] for p in pools
                  if not p["theme"] or p["theme"].startswith("[")]
    if sans_theme:
        fails.append(f"ARCH-5 : THÈME_ÉDITORIAL absent/template sur {sans_theme}")

    # ARCH-6 : crescendo QV
    qv_pools = [p for p in pools if p["type"] == "QV"]
    niveau_fails = []
    for p in qv_pools:
        attendu = niveau_attendu(p["position"])
        if attendu and p["niveau"] and p["niveau"] != attendu:
            niveau_fails.append(f"{p['pool_id']} pos={p['position']} niveau={p['niveau']} attendu={attendu}")
    if niveau_fails:
        warnings.append(f"ARCH-6 (crescendo QV) : {niveau_fails}")

    return fails, warnings

# ── ÉTAT PIPELINE ────────────────────────────────────────────────────────────

def update_pipeline_state(ligne, fichier, mtime, verdict, fails):
    state = {}
    if STATE_FILE.exists():
        try:
            state = json.loads(STATE_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass

    lignes = state.setdefault("lignes", {})
    ligne_state = lignes.setdefault(ligne, {"gates": {}})
    gates = ligne_state.setdefault("gates", {})

    gates["A4"] = {
        "status":        verdict,
        "timestamp":     datetime.now().isoformat(timespec="seconds"),
        "fichier":       str(fichier),
        "fichier_mtime": mtime,
        "fails":         fails,
    }

    # Poser ou lever le flag B2
    b2 = gates.setdefault("B2", {"status": "UNKNOWN", "timestamp": None, "fichier": None, "fichier_mtime": None, "fails": []})
    if verdict == "NO_GO":
        b2["status"] = "LOCKED"
    elif verdict == "GO" and b2["status"] == "LOCKED":
        b2["status"] = "UNKNOWN"  # déverrouillé, en attente du fichier B2

    STATE_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

def update_dashboard_state(ligne, verdict, fails):
    if not DASH_FILE.exists():
        return
    try:
        state = json.loads(DASH_FILE.read_text(encoding="utf-8"))
    except Exception:
        return

    lignes = state.setdefault("lignes", {})
    entry  = lignes.setdefault(ligne, {})

    entry["gate_a4"]      = verdict
    entry["gate_a4_fails"] = fails
    entry["bloque"]       = (verdict == "NO_GO")

    if verdict == "NO_GO":
        entry["note"] = f"⛔ Gate A4 NO_GO — B2 verrouillé ({len(fails)} fail(s))"
    else:
        entry["note"] = entry.get("note", "").replace("⛔ Gate A4 NO_GO — B2 verrouillé", "").strip() or entry.get("note", "")

    state["last_updated"] = datetime.now().strftime("%Y-%m-%d")
    DASH_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

# ── RAPPORT ──────────────────────────────────────────────────────────────────

def run(fichier, ligne):
    path = Path(fichier)
    if not path.exists():
        print(f"ERREUR : fichier introuvable — {fichier}")
        sys.exit(2)

    mtime = os.path.getmtime(path)

    # Vérifier si déjà évalué sur ce fichier (même mtime)
    if STATE_FILE.exists():
        try:
            state = json.loads(STATE_FILE.read_text(encoding="utf-8"))
            prev = state.get("lignes", {}).get(ligne, {}).get("gates", {}).get("A4", {})
            if prev.get("fichier_mtime") == mtime and prev.get("status") in ("GO", "NO_GO"):
                verdict = prev["status"]
                print(f"\nGATE A4 [{ligne}] — déjà évalué sur ce fichier")
                print(f"VERDICT : {verdict} (inchangé)\n")
                return verdict
        except Exception:
            pass

    print(f"\n{'='*60}")
    print(f"GATE A4 — {ligne}")
    print(f"Fichier : {path.name}")
    print(f"{'='*60}\n")

    pools, err = load_pools(path)
    if err:
        print(f"ERREUR lecture : {err}")
        update_pipeline_state(ligne, fichier, mtime, "NO_GO", [err])
        update_dashboard_state(ligne, "NO_GO", [err])
        sys.exit(2)

    print(f"Pools chargés : {len(pools)}\n")

    fails, warnings = run_checks(pools)

    for f in fails:
        print(f"  ✗ {f}")
    for w in warnings:
        print(f"  ⚠ {w}")

    verdict = "NO_GO" if fails else "GO"

    print(f"\n{'─'*60}")
    print(f"GATE A4 VERDICT : {verdict}")
    if verdict == "NO_GO":
        print(f"  → B2 VERROUILLÉ pour {ligne}")
    else:
        print(f"  → B2 déverrouillé pour {ligne}")
    print(f"{'='*60}\n")

    update_pipeline_state(ligne, fichier, mtime, verdict, fails)
    update_dashboard_state(ligne, verdict, fails)

    return verdict

# ── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Gate A4 — Architecture pools")
    parser.add_argument("fichier", help="Chemin vers A4_LIGNE.xlsx")
    parser.add_argument("ligne",   help="Code ligne (ex: MAYENNE, CDM...)")
    args = parser.parse_args()

    verdict = run(args.fichier, args.ligne)
    sys.exit(0 if verdict == "GO" else 1)
