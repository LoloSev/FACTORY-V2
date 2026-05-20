#!/usr/bin/env python3
"""
validate_qa_metrics.py — Gate QA métriques (densités, veille, stock)
Usage: python validate_qa_metrics.py <fichier.xlsx> [--sheet STOCK_MAITRE] [--cible N]

Checks (zéro token) :
  VEI   Marqueurs veille (liste exhaustive RULE-OBS-008)
  DEN   Densité par entité / pool (>2 occurrences du même mot-clé)
  STK   Stock actuel vs cible (--cible, défaut 277)
  DUP   Doublons exacts question↔question (même libellé normalisé)
  LEN   Distribution longueur libellés

Retourne : rapport détaillé + verdict PASS / WARNING
"""

import sys
import re
import argparse
import openpyxl
from collections import Counter, defaultdict

# ── CONFIG ─────────────────────────────────────────────────────────────────

MARQUEURS_VEILLE = {
    "dernier": 5, "dernière": 5,
    "premier": 5, "première": 5,
    "jamais": 4, "encore jamais": 4,
    "seul": 4, "unique": 4,
    "record": 4, "recordman": 4, "recordwoman": 4,
    "plus grand": 4, "plus petit": 4, "plus rapide": 4,
    "meilleur": 4, "pire": 4,
    "à ce jour": 2,
}

# Mots à ignorer pour densité (stopwords thématiques)
STOPWORDS_DENSITÉ = {
    "le", "la", "les", "un", "une", "des", "de", "du", "en", "et",
    "est", "qui", "que", "dans", "sur", "par", "pour", "avec",
    "quel", "quelle", "quels", "quelles", "à", "au", "aux",
    "comment", "combien", "quand", "où",
}

FORMATS = {
    "STOCK_MAITRE": {"q_id": "retroId", "question": "question", "answer": "answer", "pool": "poolHistorique"},
    "B5_TABLEUR":   {"q_id": "ID Global", "question": "Question", "answer": "Bonne réponse", "pool": "Pool principal"},
    "QUESTIONS":    {"q_id": "Q_ID", "question": "LIBELLÉ", "answer": "RÉPONSE", "pool": "POOL_ID"},
}

# ── LECTURE ─────────────────────────────────────────────────────────────────

def detect_format(headers):
    for fmt_name, mapping in FORMATS.items():
        if mapping["question"] in headers and mapping["answer"] in headers:
            return fmt_name, mapping
    return None, None

def load_questions(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames
    questions = []

    for sname in sheets:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = None
        mapping = None
        header_row_idx = None
        for i, row in enumerate(rows[:5]):
            row_vals = [str(v) if v is not None else "" for v in row]
            fmt_name, mapping = detect_format(row_vals)
            if mapping:
                headers = row_vals
                header_row_idx = i
                break

        if not headers:
            continue

        col = {v: headers.index(v) for v in mapping.values() if v and v in headers}

        for row in rows[header_row_idx + 1:]:
            if not any(v is not None for v in row):
                continue

            def get(field, _mapping=mapping, _col=col, _row=row):
                key = _mapping.get(field)
                if key and key in _col:
                    v = _row[_col[key]]
                    return str(v).strip() if v is not None else ""
                return ""

            q = {
                "q_id":     get("q_id") or f"row_{len(questions)+1}",
                "question": get("question"),
                "answer":   get("answer"),
                "pool":     get("pool") or sname,
                "sheet":    sname,
            }
            if q["question"] or q["answer"]:
                questions.append(q)

    wb.close()
    return questions

# ── CHECKS ──────────────────────────────────────────────────────────────────

def check_veille(questions):
    """Détecte les marqueurs d'obsolescence."""
    flagged = []
    for q in questions:
        text = q["question"].lower()
        for marqueur, type_risque in MARQUEURS_VEILLE.items():
            if marqueur in text:
                flagged.append({
                    "q_id": q["q_id"], "pool": q["pool"],
                    "marqueur": marqueur, "type": f"TYPE-{type_risque}",
                    "question": q["question"][:80],
                })
                break  # un seul flag par question
    return flagged

def check_densité(questions, seuil_entité=3, seuil_pool_pct=30):
    """Détecte surreprésentation mots-clés (entités, termes récurrents)."""
    warnings = []

    # Densité par pool
    pool_counts = Counter(q["pool"] for q in questions)
    total = len(questions)

    # Densité mots-clés dans libellés (approximation entités)
    mot_counter = Counter()
    mot_pool = defaultdict(lambda: defaultdict(int))

    for q in questions:
        mots = [m for m in re.findall(r"\b[A-ZÀ-Ûa-zà-û]{3,}\b", q["question"].lower())
                if m not in STOPWORDS_DENSITÉ]
        for m in set(mots):  # set pour éviter double-comptage intra-question
            mot_counter[m] += 1
            mot_pool[q["pool"]][m] += 1

    # Mots surreprésentés globalement
    for mot, count in mot_counter.most_common(20):
        if count >= seuil_entité:
            warnings.append(f"DEN_ENTITÉ '{mot}' apparaît dans {count} questions")

    # Pool avec très peu de questions (possible déséquilibre)
    pools_with_q = {k: v for k, v in pool_counts.items() if k and k != "None"}
    if pools_with_q:
        max_pool = max(pools_with_q.values())
        min_pool = min(pools_with_q.values())
        if max_pool > 0 and (min_pool / max_pool) < 0.3:
            warnings.append(
                f"DEN_POOL déséquilibre détecté : "
                f"pool max={max_pool} questions, pool min={min_pool} questions"
            )

    return warnings

def check_stock(questions, cible):
    """Vérifie stock total vs cible."""
    total = len(questions)
    manquant = cible - total
    return total, manquant

def check_doublons(questions):
    """Détecte libellés identiques (normalisés)."""
    seen = {}
    doublons = []
    for q in questions:
        norm = re.sub(r"\s+", " ", q["question"].lower().strip())
        if norm in seen:
            doublons.append((seen[norm], q["q_id"], q["question"][:60]))
        else:
            seen[norm] = q["q_id"]
    return doublons

def check_longueurs(questions):
    """Distribution des longueurs de libellés."""
    stats = {"<=5": 0, "6-10": 0, "11-15": 0, ">15": 0}
    for q in questions:
        n = len(q["question"].split())
        if n <= 5:
            stats["<=5"] += 1
        elif n <= 10:
            stats["6-10"] += 1
        elif n <= 15:
            stats["11-15"] += 1
        else:
            stats[">15"] += 1
    return stats

# ── RAPPORT ──────────────────────────────────────────────────────────────────

def run(path, sheet_name=None, cible=277):
    print(f"\n{'='*60}")
    print(f"QA METRICS REPORT — {path}")
    if sheet_name:
        print(f"Sheet  : {sheet_name}")
    print(f"Cible  : {cible} questions")
    print(f"{'='*60}\n")

    questions = load_questions(path, sheet_name)
    if not questions:
        print("⚠️  Aucune question trouvée.")
        sys.exit(1)

    print(f"Questions chargées : {len(questions)}\n")

    has_warnings = False

    # STOCK
    total, manquant = check_stock(questions, cible)
    stock_status = "✓" if manquant <= 0 else "⚠"
    print(f"[STOCK] {stock_status} {total}/{cible} questions")
    if manquant > 0:
        print(f"        → Déficit : {manquant} questions manquantes")
        has_warnings = True
    print()

    # DOUBLONS
    doublons = check_doublons(questions)
    print(f"[DOUBLONS] {len(doublons)} détectés")
    for (id_a, id_b, texte) in doublons:
        print(f"  ⚠ {id_a} ↔ {id_b} — '{texte}...'")
        has_warnings = True
    print()

    # VEILLE
    veille = check_veille(questions)
    print(f"[VEILLE] {len(veille)} questions à risque d'obsolescence")
    for v in veille:
        print(f"  ⚠ {v['q_id']} ({v['pool']}) [{v['type']}] marqueur='{v['marqueur']}'")
        print(f"    '{v['question']}'")
    if veille:
        has_warnings = True
    print()

    # DENSITÉ
    den_warnings = check_densité(questions)
    print(f"[DENSITÉ] {len(den_warnings)} warnings")
    for w in den_warnings:
        print(f"  ⚠ {w}")
        has_warnings = True
    print()

    # LONGUEURS
    stats = check_longueurs(questions)
    print(f"[LONGUEURS LIBELLÉS]")
    for k, v in stats.items():
        bar = "█" * (v // max(1, total // 30))
        flag = " ← ⚠ hors cible" if k in (">15", "11-15") and v > 0 else ""
        print(f"  {k:6} mots : {v:4} {bar}{flag}")
    if stats[">15"] > 0 or stats["11-15"] > 0:
        has_warnings = True
    print()

    # RÉPARTITION PAR POOL
    pool_counts = Counter(q["pool"] for q in questions)
    print(f"[RÉPARTITION PAR POOL]")
    for pool, count in sorted(pool_counts.items()):
        print(f"  {str(pool):<20} : {count} questions")
    print()

    verdict = "WARNING" if has_warnings else "PASS"
    print(f"{'─'*60}")
    print(f"QA_METRICS_GLOBAL : {verdict}")
    print(f"{'='*60}\n")

    return verdict

# ── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="QA Métriques — FACTORY")
    parser.add_argument("fichier", help="Chemin vers le fichier xlsx")
    parser.add_argument("--sheet", default=None, help="Nom de la feuille")
    parser.add_argument("--cible", type=int, default=277, help="Stock cible (défaut 277)")
    args = parser.parse_args()

    verdict = run(args.fichier, args.sheet, args.cible)
    sys.exit(0 if verdict == "PASS" else 2)  # exit 2 = warning, pas bloquant
