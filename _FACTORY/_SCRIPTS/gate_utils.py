"""
gate_utils.py — Fonctions communes à tous les scripts gate_*.py
"""
import json
import os
import sys
from pathlib import Path
from datetime import datetime

ROOT       = Path(__file__).parent.parent
STATE_FILE = ROOT / "_STATE" / "pipeline_state.json"
DASH_FILE  = ROOT / "_STATE" / "DASHBOARD_STATE.json"

# ── PIPELINE STATE ────────────────────────────────────────────────────────────

def load_state():
    if not STATE_FILE.exists():
        return {}
    try:
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}

def save_state(state):
    STATE_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

def get_prev_mtime(ligne, etape):
    state = load_state()
    try:
        return state["lignes"][ligne]["gates"][etape]["fichier_mtime"]
    except (KeyError, TypeError):
        return None

def already_evaluated(fichier, ligne, etape):
    """True si le fichier n'a pas changé depuis la dernière évaluation GO/NO_GO."""
    mtime = os.path.getmtime(fichier)
    prev  = get_prev_mtime(ligne, etape)
    if prev is None or prev != mtime:
        return False, mtime
    state = load_state()
    try:
        status = state["lignes"][ligne]["gates"][etape]["status"]
        if status in ("GO", "NO_GO"):
            return True, mtime
    except (KeyError, TypeError):
        pass
    return False, mtime

def update_gate(ligne, etape, fichier, mtime, verdict, fails,
                next_etape=None):
    """Met à jour pipeline_state.json pour une gate donnée."""
    state = load_state()
    lignes = state.setdefault("lignes", {})
    gates  = lignes.setdefault(ligne, {}).setdefault("gates", {})

    try:
        fichier_ref = str(Path(fichier).resolve().relative_to(ROOT.resolve()))
    except Exception:
        fichier_ref = str(fichier)

    gates[etape] = {
        "status":        verdict,
        "timestamp":     datetime.now().isoformat(timespec="seconds"),
        "fichier":       fichier_ref,
        "fichier_mtime": mtime,
        "fails":         fails,
    }

    # Verrouiller / déverrouiller l'étape suivante
    if next_etape:
        nxt = gates.setdefault(next_etape, {
            "status": "UNKNOWN", "timestamp": None,
            "fichier": None, "fichier_mtime": None, "fails": []
        })
        if verdict == "NO_GO":
            nxt["status"] = "LOCKED"
        elif verdict == "GO" and nxt.get("status") == "LOCKED":
            nxt["status"] = "UNKNOWN"

    save_state(state)

def update_dashboard(ligne, etape, verdict, fails):
    """Met à jour DASHBOARD_STATE.json."""
    if not DASH_FILE.exists():
        return
    try:
        state = json.loads(DASH_FILE.read_text(encoding="utf-8"))
    except Exception:
        return

    entry = state.setdefault("lignes", {}).setdefault(ligne, {})
    key   = f"gate_{etape.lower()}"
    entry[key]              = verdict
    entry[f"{key}_fails"]   = fails
    entry["bloque"]         = (verdict == "NO_GO")

    if verdict == "NO_GO":
        entry["note"] = f"⛔ Gate {etape} NO_GO ({len(fails)} fail(s))"
    state["last_updated"] = datetime.now().strftime("%Y-%m-%d")
    DASH_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

# ── AFFICHAGE ─────────────────────────────────────────────────────────────────

def header(etape, ligne, fichier):
    print(f"\n{'='*60}")
    print(f"GATE {etape} — {ligne}")
    print(f"Fichier : {Path(fichier).name}")
    print(f"{'='*60}\n")

def footer(etape, ligne, verdict, next_etape=None):
    print(f"\n{'─'*60}")
    print(f"GATE {etape} VERDICT : {verdict}")
    if next_etape:
        if verdict == "NO_GO":
            print(f"  → {next_etape} VERROUILLÉ pour {ligne}")
        else:
            print(f"  → {next_etape} déverrouillé pour {ligne}")
    print(f"{'='*60}\n")

# ── LECTURE XLSX GÉNÉRIQUE ────────────────────────────────────────────────────

def load_sheet(path, sheet_name):
    """Charge une feuille xlsx → liste de dicts. Retourne (rows, error)."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None, f"Feuille '{sheet_name}' absente"
    ws  = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], None

    headers = [str(v) if v is not None else "" for v in rows[0]]
    result  = []
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        d = {}
        for i, h in enumerate(headers):
            v = row[i] if i < len(row) else None
            d[h] = str(v).strip() if v is not None else ""
        result.append(d)
    return result, None

def check_required_fields(rows, fields, label=""):
    """Retourne liste de fails pour champs obligatoires manquants."""
    fails = []
    for r in rows:
        missing = [f for f in fields if not r.get(f)]
        if missing:
            id_val = r.get("Q_ID") or r.get("POOL_ID") or "?"
            fails.append(f"{label}champs manquants {missing} — {id_val}")
    return fails
