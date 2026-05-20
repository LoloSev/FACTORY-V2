"""
sync_glossaire.py — FACTORY
============================
Synchronise les onglets GLOSSAIRE de tous les xlsx du template
depuis la source de vérité : glossaire_documentaire_factory.md

Détection intelligente : sync uniquement si le glossaire a changé
depuis la dernière exécution. No-op silencieux sinon.

Usage :
    python sync_glossaire.py [--lignes] [--all] [--force]

Options :
    (aucune)    : synchronise uniquement _TEMPLATE si glossaire modifié
    --lignes    : synchronise aussi toutes les lignes actives
    --all       : idem --lignes
    --force     : force la sync même si glossaire inchangé

Date : 2026-05-18
"""

import sys
import re
import os
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --- Config chemins ---
FACTORY_ROOT = Path(__file__).parent.parent
GLOSSAIRE    = FACTORY_ROOT / "_STANDARDS" / "_GLOBAL" / "glossaire_documentaire_factory.md"
TEMPLATE_DIR = FACTORY_ROOT / "_LIGNES" / "_TEMPLATE"
LIGNES_DIR   = FACTORY_ROOT / "_LIGNES"

STEP_FILES = {
    'A2':     "A2_APPRO/A2_THEME.xlsx",
    'A3':     "A3_TRAITEMENT/A3_THEME.xlsx",
    'A4':     "A4_POOLS/A4_THEME.xlsx",
    'B2':     "B2_GENERATION/B2_THEME.xlsx",
    'B3':     "B3_DISTRACTEURS/B3_THEME.xlsx",
    'B5':     "B5_AUDIT/B5_THEME.xlsx",
    'EXPORT': "EXPORT/QUIZ_THEME_EXPORT.xlsx",
}

# Termes par étape — source de configuration
STEP_TERMS = {
    'A2': ['BIB', 'ACCESSIBILITE_GRAND_PUBLIC', 'COBAYE', 'FACTORY_PIPELINE',
           'NAMING_PATTERN_ARTIFACT', 'NAMING_PATTERN_PROCESS_DOC',
           'TOKEN_OPTIMIZATION_MONITORING', 'SESSION_RESUMPTION_PROMPT',
           'STATUS_STANDARD', 'DEPENDENCY', 'MACHINE_READABLE'],
    'A3': ['ANGLE_DOCUMENTAIRE', 'ANGLE_ASSIGNMENT', 'BIPREGEN', 'ANGIPREGEN',
           'PROCESS_BIB', 'COHERENCE_CULTURELLE', 'ACCESSIBILITE_GRAND_PUBLIC',
           'OBJET_DOCUMENTAIRE', 'STATUS_STANDARD', 'MACHINE_READABLE',
           'TOKEN_OPTIMIZATION_MONITORING', 'NAMING_PATTERN_PROCESS_DOC'],
    'A4': ['POOL', 'IF_SF', 'IF_ROT', 'QV', 'POOLS_ARCHITECTURE', 'POOLS_DOC',
           'POOL_COLLISION', 'ROTATION_LENTE', 'ROTATION_RAPIDE', 'FORTE_VARIETE',
           'ONBOARDING_FLUIDE', 'ANGLE_ASSIGNMENT', 'COHERENCE_CULTURELLE',
           'STATUS_STANDARD', 'MACHINE_READABLE', 'TOKEN_OPTIMIZATION_MONITORING'],
    'B2': ['POOL', 'IF_SF', 'IF_ROT', 'QV', 'POOL_COLLISION',
           'ONBOARDING_FLUIDE', 'COHERENCE_CULTURELLE', 'ANGLE_DOCUMENTAIRE',
           'ANGLE_ASSIGNMENT', 'STATUS_STANDARD', 'TOKEN_OPTIMIZATION_MONITORING',
           'DECISION_GATE'],
    'B3': ['DISTRACTEUR', 'POOL_COLLISION', 'HARD_COLLISION', 'SOFT_COLLISION',
           'PASS_FUNNEL', 'DECISION_GATE', 'PLAUSIBILITY_RATING',
           'FORMAT_HOMOGENEITY', 'REUSE_RATE', 'BIAS_DETECTION',
           'STATUS_STANDARD', 'TOKEN_OPTIMIZATION_MONITORING', 'QA_STATUS'],
    'B5': ['QA_STATUS', 'AUDIT_VALIDATION', 'POOL_COLLISION',
           'COHERENCE_CULTURELLE', 'HARD_COLLISION', 'SOFT_COLLISION',
           'STATUS_STANDARD', 'TOKEN_OPTIMIZATION_MONITORING', 'RULES_EXTRACTION'],
    'EXPORT': ['QA_STATUS', 'SUFFIXES_ETAT_XLSX', 'STATUS_STANDARD',
               'FACTORY_PIPELINE', 'NAMING_PATTERN_ARTIFACT'],
}

# --- Styles ---
def _h(c): c.font=Font(bold=True,color="FFFFFF",name="Arial",size=10); c.fill=PatternFill("solid",start_color="1F4E79"); c.alignment=Alignment(horizontal="center",vertical="center")
def _s(c): c.font=Font(bold=True,color="1F4E79",name="Arial",size=9); c.fill=PatternFill("solid",start_color="D6E4F0"); c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
def _w(c): c.font=Font(color="7F6000",name="Arial",size=9,italic=True); c.fill=PatternFill("solid",start_color="FFF2CC"); c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
def _n(c): c.font=Font(name="Arial",size=9); c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
def _k(c): c.font=Font(name="Arial",size=8,color="595959",italic=True); c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)

# --- Parser glossaire ---
def parse_glossaire(path):
    terms = {}
    content = Path(path).read_text(encoding="utf-8")
    sections = re.split(r'\n# SECTION — ', content)
    for sec in sections[1:]:
        lines = sec.strip().split('\n')
        defs = re.findall(r'\[DEF-[\w-]+\]\n(\w[\w_]+):\n(.*?)(?=\n\[|\n# |\Z)', sec, re.DOTALL)
        rules = re.findall(r'\[RULE-([\w-]+)\]\n(.+?)(?=\n\[|\n# |\Z)', sec, re.DOTALL)
        for name, body in defs:
            short = body.strip().split('\n')[0][:120]
            terms[name] = {
                'definition': short,
                'rules': [f"[RULE-{r[0]}] {r[1].strip().split(chr(10))[0][:100]}" for r in rules[:2]]
            }
    return terms

# --- Écrire onglet GLOSSAIRE ---
def write_glossaire_tab(wb, step, terms_db):
    if 'GLOSSAIRE' in wb.sheetnames:
        del wb['GLOSSAIRE']
    ws = wb.create_sheet('GLOSSAIRE')

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 55

    r = 1
    t = ws.cell(r,1, f"GLOSSAIRE — ÉTAPE {step}"); t.font=Font(bold=True,color="1F4E79",name="Arial",size=12); ws.merge_cells(f"A{r}:C{r}"); r+=1
    t = ws.cell(r,1,"Source : glossaire_documentaire_factory.md — LECTURE SEULE"); _k(t); ws.merge_cells(f"A{r}:C{r}"); r+=2

    for col,val,w in [(1,"TERME",28),(2,"DÉFINITION COURTE",60),(3,"RÈGLE PRINCIPALE",55)]:
        c=ws.cell(r,col,val); _h(c); ws.column_dimensions[get_column_letter(col)].width=w
    ws.freeze_panes=f"A{r+1}"; r+=1

    step_keys = STEP_TERMS.get(step, [])
    found, missing_keys = [], []
    for key in step_keys:
        if key in terms_db:
            t = terms_db[key]
            _n(ws.cell(r,1,key)); _n(ws.cell(r,2,t['definition'])); _n(ws.cell(r,3,t['rules'][0] if t['rules'] else "")); r+=1
            found.append(key)
        else:
            _n(ws.cell(r,1,key)); _n(ws.cell(r,2,"[non trouvé dans glossaire — vérifier source]")); _n(ws.cell(r,3,"")); r+=1
            missing_keys.append(key)

    r+=1
    c=ws.cell(r,1,"⚠ ALERTES — TERMES POTENTIELLEMENT MANQUANTS"); _s(c); ws.merge_cells(f"A{r}:C{r}"); r+=1
    c=ws.cell(r,1,"L'IA consigne ici tout terme utilisé dans cette étape mais absent du glossaire source."); _k(c); ws.merge_cells(f"A{r}:C{r}"); r+=1

    # Alertes sur clés non trouvées
    alerted = False
    for key in missing_keys:
        _w(ws.cell(r,1,"MANQUANT")); _w(ws.cell(r,2,f"{key} — présent dans config step mais absent du glossaire")); _w(ws.cell(r,3,"→ Ajouter dans glossaire_documentaire_factory.md")); r+=1
        alerted = True

    if not alerted:
        _n(ws.cell(r,1,"Aucune alerte automatique — vérification manuelle recommandée à chaque étape")); r+=1

    r+=1
    c=ws.cell(r,1,f"sync_glossaire.py — dernière exécution automatique — {__import__('datetime').date.today()}"); _k(c); ws.merge_cells(f"A{r}:C{r}")

    return len(found), len(missing_keys)

# --- Main ---
def sync_dir(base_dir, theme="THEME"):
    terms_db = parse_glossaire(GLOSSAIRE)
    print(f"\n{'─'*50}")
    print(f"Sync → {base_dir.name}")
    print(f"Glossaire : {len(terms_db)} termes parsés")
    print(f"{'─'*50}")

    for step, rel_path in STEP_FILES.items():
        # Adapter le nom si pas TEMPLATE
        actual_path = base_dir / rel_path.replace("THEME", theme)
        if not actual_path.exists():
            # Essayer avec THEME littéral
            actual_path = base_dir / rel_path
        if not actual_path.exists():
            print(f"  {step} — IGNORÉ (fichier absent : {rel_path})")
            continue
        try:
            wb = load_workbook(actual_path)
        except (FileNotFoundError, OSError):
            print(f"  {step} — IGNORÉ (ghost inode : {actual_path.name})")
            continue
        found, missing = write_glossaire_tab(wb, step, terms_db)
        wb.save(actual_path)
        status = "✓" if missing == 0 else f"⚠ {missing} clé(s) manquante(s)"
        print(f"  {step} — {found} termes sync {status}")

# --- Détection de changement ---
STAMP_FILE = FACTORY_ROOT / "_STATE" / ".sync_stamp.json"

def get_glossaire_mtime():
    return os.path.getmtime(GLOSSAIRE)

def load_stamp():
    if STAMP_FILE.exists():
        try:
            return json.loads(STAMP_FILE.read_text())
        except Exception:
            return {}
    return {}

def save_stamp(mtime):
    import datetime
    STAMP_FILE.write_text(json.dumps({
        "glossaire_mtime": mtime,
        "last_sync": str(datetime.datetime.now())
    }))

def needs_sync(force=False):
    import datetime
    if force:
        return True, "force"
    stamp = load_stamp()
    current_mtime = get_glossaire_mtime()
    last_mtime = stamp.get("glossaire_mtime", 0)
    if current_mtime > last_mtime:
        import datetime as dt
        ts = dt.datetime.fromtimestamp(current_mtime).strftime('%Y-%m-%d %H:%M')
        return True, f"glossaire modifie ({ts})"
    return False, f"glossaire inchange depuis {stamp.get('last_sync', 'inconnu')}"

if __name__ == "__main__":
    sync_lignes = '--lignes' in sys.argv or '--all' in sys.argv
    force       = '--force' in sys.argv

    do_sync, reason = needs_sync(force)

    if not do_sync:
        print(f"OK GLOSSAIRE SYNC -- no-op ({reason})")
        sys.exit(0)

    print(f"SYNC GLOSSAIRE -- {reason}")

    sync_dir(TEMPLATE_DIR)

    if sync_lignes:
        for ligne_dir in sorted(LIGNES_DIR.iterdir()):
            if ligne_dir.is_dir() and ligne_dir.name.startswith('_') and ligne_dir.name != '_TEMPLATE':
                theme = ligne_dir.name.lstrip('_')
                sync_dir(ligne_dir, theme)

    save_stamp(get_glossaire_mtime())
    print()
    print("Sync termine. Stamp mis a jour.")
    print("--force pour forcer | --lignes pour toutes les lignes")
