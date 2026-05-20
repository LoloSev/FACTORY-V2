#!/usr/bin/env python3
"""
validate_collision.py — Gate COLLISION (anti-collision inter-pools)
Usage: python validate_collision.py <fichier.xlsx> [--sheet STOCK_MAITRE]

Checks (zéro token) :
  COL-1  Réponse de Q_A = réponse de Q_B (même texte, pools différents)
  COL-2  Réponse de Q_A contenue dans le libellé de Q_B (autre pool)
  COL-3  Doublon intra-pool : même réponse + même pool
  COL-D1 Distracteur = réponse correcte ailleurs (si distracteurs disponibles)
  COL-D2 Unicité intra-question : D1=D2 ou D1=D3 ou D2=D3

Retourne : GO / NO_GO + rapport détaillé
"""

import sys
import re
import argparse
import openpyxl
from collections import defaultdict

# ── MAPPING COLONNES ────────────────────────────────────────────────────────

FORMATS = {
    "STOCK_MAITRE": {
        "q_id": "retroId", "question": "question", "answer": "answer",
        "pool": "poolHistorique", "d1": None, "d2": None, "d3": None
    },
    "B5_TABLEUR": {
        "q_id": "ID Global", "question": "Question", "answer": "Bonne réponse",
        "pool": "Pool principal",
        "d1": "Distracteurs possibles", "d2": None, "d3": None  # colonne unique pour distracteurs
    },
    "QUESTIONS": {
        "q_id": "Q_ID", "question": "LIBELLÉ", "answer": "RÉPONSE",
        "pool": "POOL_ID", "d1": "D1", "d2": "D2", "d3": "D3"
    },
}

# ── LECTURE ─────────────────────────────────────────────────────────────────

def detect_format(headers):
    for fmt_name, mapping in FORMATS.items():
        if mapping["question"] in headers and mapping["answer"] in headers:
            return fmt_name, mapping
    return None, None

def load_all_questions(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames
    questions = []

    for sname in sheets:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = None
        mapping = None
        header_row_idx = None
        for i, row in enumerate(rows[:5]):
            row_vals = [str(v) if v is not None else "" for v in row]
            fmt_name, mapping = detect_format(row_vals)
            if mapping:
                headers = row_vals
                header_row_idx = i
                break

        if not headers:
            continue

        col = {v: headers.index(v) for v in mapping.values() if v and v in headers}

        for row in rows[header_row_idx + 1:]:
            if not any(v is not None for v in row):
                continue

            def get(field):
                key = mapping.get(field)
                if key and key in col:
                    v = row[col[key]]
                    return str(v).strip() if v is not None else ""
                return ""

            questions.append({
                "q_id":     get("q_id") or f"row_{len(questions)+1}",
                "question": get("question"),
                "answer":   get("answer"),
                "pool":     get("pool") or sname,
                "d1":       get("d1"),
                "d2":       get("d2"),
                "d3":       get("d3"),
                "sheet":    sname,
            })

    wb.close()
    return questions

# ── NORMALISATION ────────────────────────────────────────────────────────────

def normalize(text):
    """Normalise pour comparaison : minuscules, sans accents grossiers."""
    if not text:
        return ""
    t = text.lower().strip()
    t = re.sub(r"\s+", " ", t)
    return t

# ── CHECKS ──────────────────────────────────────────────────────────────────

def check_collisions(questions):
    hard = []
    soft = []

    # Index : réponse normalisée → liste de (q_id, pool)
    answer_index = defaultdict(list)
    for q in questions:
        if q["answer"]:
            answer_index[normalize(q["answer"])].append((q["q_id"], q["pool"]))

    # Index distracteurs : valeur normalisée → liste de (q_id, pool, position)
    distractor_index = defaultdict(list)
    for q in questions:
        for pos, field in [("D1", "d1"), ("D2", "d2"), ("D3", "d3")]:
            val = q.get(field, "")
            if val:
                distractor_index[normalize(val)].append((q["q_id"], q["pool"], pos))

    checked_pairs = set()

    for q in questions:
        ans_norm = normalize(q["answer"])
        q_text_norm = normalize(q["question"])

        # COL-1 : même réponse dans deux pools différents
        if ans_norm and ans_norm in answer_index:
            for (other_id, other_pool) in answer_index[ans_norm]:
                if other_id == q["q_id"]:
                    continue
                pair = tuple(sorted([q["q_id"], other_id]))
                if pair in checked_pairs:
                    continue
                checked_pairs.add(pair)
                if other_pool != q["pool"]:
                    hard.append({
                        "type": "COL-1_RÉPONSE_IDENTIQUE_INTER_POOLS",
                        "q_a": q["q_id"], "pool_a": q["pool"],
                        "q_b": other_id, "pool_b": other_pool,
                        "valeur": q["answer"],
                    })
                else:
                    # COL-3 : doublon intra-pool
                    soft.append({
                        "type": "COL-3_DOUBLON_INTRA_POOL",
                        "q_a": q["q_id"], "q_b": other_id,
                        "pool": q["pool"], "valeur": q["answer"],
                    })

        # COL-2 : réponse de Q_A dans le libellé de Q_B (autre pool)
        if ans_norm and len(ans_norm) > 2:
            for other_q in questions:
                if other_q["q_id"] == q["q_id"]:
                    continue
                if other_q["pool"] == q["pool"]:
                    continue
                pair = (q["q_id"], other_q["q_id"], "COL2")
                if pair in checked_pairs:
                    continue
                if ans_norm in normalize(other_q["question"]):
                    checked_pairs.add(pair)
                    soft.append({
                        "type": "COL-2_RÉPONSE_DANS_LIBELLÉ_AUTRE_POOL",
                        "q_a": q["q_id"], "pool_a": q["pool"], "valeur": q["answer"],
                        "q_b": other_q["q_id"], "pool_b": other_q["pool"],
                    })

        # COL-D1 : distracteur = réponse correcte ailleurs
        if ans_norm and ans_norm in distractor_index:
            for (dist_q_id, dist_pool, pos) in distractor_index[ans_norm]:
                if dist_q_id == q["q_id"]:
                    continue
                pair = (q["q_id"], dist_q_id, "COLD1")
                if pair in checked_pairs:
                    continue
                checked_pairs.add(pair)
                hard.append({
                    "type": "COL-D1_DISTRACTEUR_RÉPONSE_AILLEURS",
                    "q_a": q["q_id"], "pool_a": q["pool"], "valeur": q["answer"],
                    "q_b": dist_q_id, "pool_b": dist_pool, "position": pos,
                })

    # COL-D2 : unicité intra-question
    for q in questions:
        distractors = [d for d in [q.get("d1"), q.get("d2"), q.get("d3")] if d]
        if len(distractors) != len(set([normalize(d) for d in distractors])):
            hard.append({
                "type": "COL-D2_DISTRACTEURS_NON_UNIQUES",
                "q_id": q["q_id"], "pool": q["pool"],
                "distracteurs": distractors,
            })

    return hard, soft

# ── RAPPORT ──────────────────────────────────────────────────────────────────

def run(path, sheet_name=None):
    print(f"\n{'='*60}")
    print(f"COLLISION REPORT — {path}")
    if sheet_name:
        print(f"Sheet : {sheet_name}")
    print(f"{'='*60}\n")

    questions = load_all_questions(path, sheet_name)
    # Filtrer questions sans réponse
    questions = [q for q in questions if q["question"] or q["answer"]]

    if not questions:
        print("⚠️  Aucune question trouvée.")
        sys.exit(1)

    print(f"Questions chargées : {len(questions)}\n")

    hard, soft = check_collisions(questions)

    print(f"HARD COLLISIONS : {len(hard)}")
    for h in hard:
        t = h["type"]
        if "RÉPONSE_IDENTIQUE" in t:
            print(f"  ✗ [{t}]")
            print(f"    {h['q_a']} (Pool {h['pool_a']}) ↔ {h['q_b']} (Pool {h['pool_b']}) — '{h['valeur']}'")
        elif "DISTRACTEUR_RÉPONSE" in t:
            print(f"  ✗ [{t}]")
            print(f"    Réponse de {h['q_a']} (Pool {h['pool_a']}) = {h['position']} de {h['q_b']} (Pool {h['pool_b']}) — '{h['valeur']}'")
        elif "NON_UNIQUES" in t:
            print(f"  ✗ [{t}] {h['q_id']} (Pool {h['pool']}) — {h['distracteurs']}")
        else:
            print(f"  ✗ [{t}] {h}")

    print(f"\nSOFT COLLISIONS : {len(soft)}")
    for s in soft:
        t = s["type"]
        if "DOUBLON_INTRA" in t:
            print(f"  ⚠ [{t}] {s['q_a']} ↔ {s['q_b']} (Pool {s['pool']}) — '{s['valeur']}'")
        elif "LIBELLÉ_AUTRE" in t:
            print(f"  ⚠ [{t}]")
            print(f"    Réponse '{s['valeur']}' de {s['q_a']} dans libellé de {s['q_b']} (Pool {s['pool_b']})")

    has_distractors = any(q.get("d1") or q.get("d2") or q.get("d3") for q in questions)
    if not has_distractors:
        print("\n  ℹ  Distracteurs absents — COL-D1/D2 non évalués")

    verdict = "NO_GO" if hard else "GO"
    print(f"\n{'─'*60}")
    print(f"COLLISION_GLOBAL : {verdict}")
    print(f"  HARD : {len(hard)}")
    print(f"  SOFT : {len(soft)}")
    if hard:
        hard_ids = list({h.get("q_a") or h.get("q_id") for h in hard})
        print(f"  Q_IDS_BLOQUÉES : {hard_ids}")
    print(f"{'='*60}\n")

    return verdict

# ── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Validation COLLISION — FACTORY")
    parser.add_argument("fichier", help="Chemin vers le fichier xlsx")
    parser.add_argument("--sheet", default=None, help="Nom de la feuille (défaut: toutes)")
    args = parser.parse_args()

    verdict = run(args.fichier, args.sheet)
    sys.exit(0 if verdict == "GO" else 1)
